import os
import re
import csv as csvmodule
import math
from openpyxl import load_workbook
import codecs
import xlsxwriter

def rows_to_xlsx(rows = [], fn = "test.xlsx"):
    workbook = xlsxwriter.Workbook(fn)
    worksheet = workbook.add_worksheet()

    for (i, row) in enumerate(rows):
        for (j, cell) in enumerate(row):
            worksheet.write(i, j, cell) 
    workbook.close()
    
def xlsx_to_ds(fn = "Grambank_sheet_with_categories.xlsx", toprow = None, sheet = None):
    rows = xlsx_to_rows(fn, sheet = sheet)
    if toprow:
        return [dict(zip(toprow, row)) for row in rows]
    return [dict(zip(rows[0], row)) for row in rows[1:]]

def xlsx_to_rows(fn = "Grambank_sheet_with_categories.xlsx", sheet = None):
    def strf(x):
        if x == None:
            y = u""
        elif type(x) == type(0.0):
            y = str(int(x)) if int(x) == float(x) else str(float(x))
        else:
            y = str(x)
        return y.strip()
    wb = load_workbook(fn, data_only=True)
    ws = wb.active if not sheet else wb.get_sheet_by_name(sheet)
    rows = [[strf(c.value) for c in row] for row in ws.rows]
    return rows

def d_to_coltxt(r, key = "GBID", order = []):
    def vdump(vs, key, value):
        return u''.join([u"%s: %s\n" % (k, v) for (k, v) in ([(x, vs[x]) for x in order if x in vs] + [(key, value)] + list(sorted([(k, v) for (k, v) in vs.items() if k not in [key] + order])))])
    txt = u"\n\n\n\n".join([vdump(vs, key, k) for (k, vs) in sorted(r.items())])
    return txt

def coltxt_to_d(txt, key = "GBID"):
    def remcom(txt):
        return '\n'.join([row for row in txt.split("\n") if not row.startswith("#")])
    def unchunk(chunk):
        pairs = [(takeuntil(row, ":").strip(), takeafter(row, ":").strip()) for row in chunk.split("\n") if row.strip()]
        (_, keyv) = pairs[0]
        vs = dict(pairs)
        if key in vs:
            keyv = vs[key]
        return (keyv, vs)
    return dict([unchunk(chunk) for chunk in remcom(txt).split("\n\n\n\n") if chunk.strip()])

def recdir(dirname):
    r = {}
    for f in os.listdir(dirname):
        if os.path.isdir(os.path.join(dirname, f)):
            try:
                for x in recdir(os.path.join(dirname, f)):
                    r[os.path.join(f, x)] = None
            except WindowsError:
                continue
        else:
            r[f] = None
    return r.keys()

def rnd(c, dpi = 0.125):
    if type(c) == type(()):
        return tuple([rnd(z, dpi) for z in c])
    if type(c) == type([]):
        return [rnd(z, dpi) for z in c]
    return dpi*int(c/dpi)

def entropy(pd):
    s = 0
    for (k, p) in pd.items():
        if p > 0:
            s += p*math.log(p, 2)
    return -s

def leaves(d):
    if not d:
        return set()
    l = set([k for (k, v) in d.items() if not v])
    return l.union(setall([leaves(v) for (k, v) in d.items() if v]))

from functools import reduce
def mulp(ps):
    return reduce(lambda x, y: x*y, ps, 1)

def percent(p, n = 1):
    return (("{\\bf %." + ("%sf" % n)) % (p*100)) + "\\%}"

def fmt(n, p):
    if type(n) == type(""):
        return " %s & %s " % (n, p)
    elif type(n) == type(0):
        return ("%d & " % n) + ("{\\bf %3f" % (p*100))[:-5] + "\\%}"
    else:
        return ("%.2f" % n) + " & " + ("{\\bf %.2f" % (p*100)) + "\\%}"
    
def tabular_p(d):
    r = "\\begin{tabular}{l|r r}\\hline\n"
    for (k, p) in sorted(norm(d).items(), key = lambda xy: (xy[1], xy[0]), reverse = True):
        r += str(k) + " & " + fmt(d[k], p) + "\\\\\n"
    return r + "\\end{tabular}"

#(dim1, dim2), n
def mtabular(d):
    xs = grp2l([(x, v) for ((x, y), v) in d.items()])
    ys = grp2l([(y, v) for ((x, y), v) in d.items()])
    xt = opv(xs, sum)
    yt = opv(ys, sum)
    xl = tuple(sorted(xs))
    yl = tuple(sorted(ys))
    np = opv(opv(grp2([(x, (y, v)) for ((x, y), v) in d.items()]), dict), norm)
    tr = [("",) + tuple(["\\multicolumn{2}{c}{%s}" % x for x in xl])]
    mr = [(y,) + tuple([fmt(d.get((x, y), 0), np[x].get(y, 0.0)) for x in xl]) + (yt[y],) for y in yl]
    fr = [("",) + tuple([str(xt[x]) + " & " for x in xl]) + (sum(d.values()),)]
    return quicktabular(tr + mr + fr)

def si(xs, i = 0):
    return [x[i] for x in xs]

def sdl(d, k, v):
    if k in d:
        d[k][len(d[k])] = v
    else:
        d[k] = {0: v}

def grp(l):
    r = {}
    for x in l:
        sdl(r, x[0], x[1:])
    return {k: vs.values() for (k, vs) in r.items()}

def spearman(xy):
    rho = spearman_rho(xy)
    rhos = {}
    for i in range(1000):
        xys = [(x, y) for (x, y) in xy]
        random.shuffle(xys)
        rhos[i] = spearman_rho([(x1, y2) for ((x1, y1), (x2, y2)) in zip(xy, xys)])
    return (rho, len([1 for r in rhos.values() if rho > r]))
    
#xy = [(a, b), ...]
def spearman_rho(xy):
    """ 
    rho for tied ranks, checked by comparison with Pycluster

    >>> x = [2, 8, 5, 4, 2, 6, 1, 4, 5, 7, 4]
    >>> y = [3, 9, 4, 3, 1, 7, 2, 5, 6, 8, 3]
    >>> spearman_rho_tr(x, y)
    0.9348938334114621
    """
    def _rank(m):
        (ivec, svec) = zip(*sorted(list(enumerate(m)), key=lambda x: x[1]))
        sumranks = 0
        dupcount = 0
        newlist = [0] * len(m)
        for i in range(len(m)):
            sumranks += i
            dupcount += 1
            if i == len(m) - 1 or svec[i] != svec[i + 1]:
                averank = sumranks / float(dupcount) + 1
                for j in range(i - dupcount + 1, i + 1):
                    newlist[ivec[j]] = averank
                sumranks = 0
                dupcount = 0
        return newlist
    
    m = [x for (x, y) in xy]
    n = [y for (x, y) in xy]
    m = _rank(m)
    n = _rank(n)
    num = 0.
    den_m = 0.
    den_n = 0.
    m_mean = avg(m)
    n_mean = avg(n) 
    for (i, j) in zip(m, n):
        i = i - m_mean
        j = j - n_mean
        num += i * j 
        den_m += i ** 2
        den_n += j ** 2
    return num / ((den_m * den_n) ** (0.5))


def grp2l(l):
    r = {}
    for (a, b) in l:
        r[a] = r.get(a, [])
        r[a].append(b)
    return r

def grp2(l):
    r = {}
    for (a, b) in l:
        setd(r, a, b)
    return {k: list(vs.keys()) for (k, vs) in r.items()}

def grp2fd(l):
    r = {}
    for (a, b) in l:
        if a in r:
            r[a][b] = r[a].get(b, 0) + 1
        else:
            r[a] = {b: 1}
    return r

def grp2sum(l):
    r = {}
    for (a, b) in l:
        r[a] = r.get(a, 0) + b
    return r


def stripd(d):
    return {k: v for (k, v) in d.items() if v}

def takeuntil(s, q, plus = 0):
    i = s.find(q)
    if i >= 0:
        return s[:i+plus]
    return s

def takeafter(s, q):
    i = s.find(q)
    if i >= 0:
        return s[i+len(q):]
    return s


def addfield(d, ab):
    (a, b) = ab
    d[a] = b
    return d

def setall(xs):
    a = set()
    for x in xs:
        a.update(x)
    return a

def flatten(xs):
    def xpand(xp):
        if type([]) == type(xp) or type(()) == type(xp):
            return xp
        return [xp]
    return [x for xp in xs for x in xpand(xp)]
    
def intersectall(xs):
    a = set(xs[0])
    for x in xs[1:]:
        a.intersection_update(x)
    return a

def pairs(xs):
    return [(x, y) for x in xs for y in xs if x < y]

def inv11(d):
    return dict([(v, k) for (k, v) in d.items()])
    
def inv(d):
    r = {}
    for (k, v) in d.items():
        if v in r:
            r[v].add(k)
        else:
            r[v] = set([k])
    return r

def s2(d, reverse=True):
    dt = [(b, a) for (a, b) in d.items()]
    return [(a, d[a]) for (b, a) in sorted(dt, reverse=reverse)]

def median(xxs):
    xs = list(xxs)
    xs.sort()
    return xs[len(xs) // 2]

def variance(xs):
    m = avg(xs)
    return avg([(x-m)**2 for x in xs])

def sdev(xs):
    m = avg(xs)
    return avg([(x-m)**2 for x in xs])**(0.5)

def avgxy(xs):
    return tuple([avg(vs) for vs in zip(*xs)])

def avg(xs):
    return sum(xs)/float(len(xs))

def by(xs, n = 2):
    return [tuple(xs[i:i+n]) for i in range(0, len(xs), n)]

def cleancell(x):
    while x.startswith('"'):
        x = x[1:]
    while x.endswith('"'):
        x = x[:-1]
    return x.strip()

renewline = re.compile("[\\n\\r]")
def ptab(fn, i=1, spl = "\t", clean = cleancell):
    lines = renewline.split(load(fn))[i:]
    li = [tuple([cleancell(x) for x in l.split(spl)]) for l in lines if l != ""]
    return li

def ptabu(fn, i=1, spl = "\t", enc = "utf-8-sig"):
    lines = renewline.split(loadunicode(fn, encoding = enc))[i:]
    li = [tuple([x.strip() for x in l.split(spl)]) for l in lines if l != ""]
    return li

def ptabd(fn, spl = "\t"):
    ll = ptab(fn, i = 0, spl = spl)
    return dict([(l[0], dict(zip(ll[0][1:], l[1:]))) for l in ll[1:]])

def ptabdu(fn, spl = "\t"):
    ll = ptabu(fn, i = 0, spl = spl)
    return dict([(l[0], dict(zip(unquote(ll[0][1:]), unquote(l[1:])))) for l in ll[1:]])

def ptabcsv(fn = "glottolog_signlanguages.csv", delimiter = ","):
    with open(fn, encoding = "utf-8") as csvfile: #rb or r?
        spamreader = csvmodule.DictReader(csvfile, delimiter=delimiter, quotechar='"')
        xs = [d for d in spamreader]
    return xs

csv_to_ds = ptabcsv

def ds_to_csv(ds, fn = "test.csv", delimiter = ",", encoding = "utf-8", fieldnames = []):
    if not fieldnames:
        fieldnames = list(reversed(sorted(set([k for d in ds for k in d.keys()])))) 
    with open(fn, 'w', encoding = encoding) as csvfile: #rb or r?
        spamwriter = csvmodule.DictWriter(csvfile, delimiter=delimiter, quotechar='"', fieldnames = fieldnames)
        spamwriter.writeheader()
        for d in ds:
            spamwriter.writerow(d)
    return

def unquote(xs, quote_char = '"'):
    return [x[len(quote_char):-len(quote_char)] if x.startswith(quote_char) and x.endswith(quote_char) else x for x in xs]


reline = re.compile("[\\n\\r]")
def dtab(fn = "sails_neele.tab", encoding = "utf-8", refield = re.compile("\\t"), quote_char = '"'):
    lines = reline.split(loadunicode(fn, encoding = encoding))
    lp = [unquote([x.strip() for x in refield.split(l)], quote_char) for l in lines if l.strip()]
    topline = unquote(lp[0], quote_char)
    lpd = [dict(zip(topline, l)) for l in lp[1:]]
    return lpd

def load(fn):
    f = open(fn, 'r')
    a = f.read()
    f.close()
    if a.startswith('# -*- coding: utf-8 -*-\n') or a.startswith('\xef\xbb\xbf# -*- coding: utf-8 -*-\n'):
        return loadunicode(fn)
    return a

import io
def loadunicode(fn, encoding = "utf-8"):
    with open(fn, 'r', encoding = encoding) as f:
        utxt = f.read()
    if utxt.startswith(u'\ufeff'):
        utxt = utxt[1:]
    if utxt.startswith(u'# -*- coding: utf-8 -*-\n'):
        utxt = utxt[len(u'# -*- coding: utf-8 -*-\n'):]
    return utxt

def loadu(fn, encoding = "utf-8"):
    with open(fn, 'r', encoding = encoding, errors = "replace") as f:
        utxt = f.read()
    if utxt.startswith(u'\ufeff'):
        utxt = utxt[1:]
    if utxt.startswith(u'# -*- coding: utf-8 -*-\n'):
        utxt = utxt[len(u'# -*- coding: utf-8 -*-\n'):]
    return utxt



def loadb(fn):
    f = io.open(fn, 'rb')
    a = f.read()
    f.close()
    return a

def sumds(ds, f = sum):
    return {k: sum(vs) for (k, vs) in grp2l([(i, v) for d in ds for (i, v) in d.items()]).items()}

import pickle
def loadpkl(fn):
    f = open(fn, 'r')
    a = pickle.load(f)
    f.close()
    return a

def savepkl(o, fn):
    f = open(fn, 'w')
    pickle.dump(o, f)
    f.close()
    return

def loadbpkl(fn):
    f = open(fn, 'rb')
    a = pickle.load(f)
    f.close()
    return a

def savebpkl(o, fn):
    f = open(fn, 'wb')
    pickle.dump(o, f)
    f.close()
    return

import json
def savejson(o, fn):
    savu(json.dumps(o), fn, encoding = "utf-8")
    return

def loadjson(fn):
    return json.loads(loadu(fn))

def sav(txt, fn):
    if type(txt) == type(u""):
        f = codecs.open(fn, 'w', encoding = "utf-8")
    else:
        f = open(fn, 'w')
    f.write(txt)
    f.close()
    return

def savb(txt, fn):
    f = open(fn, 'wb')
    f.write(txt)
    f.close()
    return

def savu(txt, fn, encoding = "utf-8-sig", ucc = False):
    f = codecs.open(fn, 'w', encoding) #encoding = "utf-8", 
    f.write(("# -*- coding: utf-8 -*-\n" if ucc else "") + txt)
    f.close()
    return

def restr(d, k):
    return dict([(x, d[x]) for x in k if x in d])

def setd(ds, k1, k2, v = None):
    if k1 in ds:
        ds[k1][k2] = v
    else:
        ds[k1] = {k2: v}
    return

def setd3(ds, k1, k2, k3, v = None):
    if k1 in ds:
        if k2 in ds[k1]:
            ds[k1][k2][k3] = v
        else:
            ds[k1][k2] = {k3: v}
    else:
        ds[k1] = {k2: {k3: v}}
    return


def norm(d):
    z = float(sum(d.values()))
    if z == 0.0:
        return {k: 0.0 for (k, _) in d.items()}
    return {k: x/z for (k, x) in d.items()}

#def tok(txt, spl = re.compile('\s+|\d+|([\.!\\:\;\?\&,`\(\)\[\]"])')): aven uppochner fragetecken och utropstecken och guillemotleft/right
#    toks = spl.split(txt)
#    return filter(lambda x: x not in [None, ''], toks)
def tok(txt, spl = re.compile('\s+|\d+|([\.!\\:\;\?\&,`\(\)\[\]"])')):
    toks = spl.split(txt)
    return filter(lambda x: x not in [None, ''], toks)


retag = re.compile("<[^\>\<]+?>")
rerem = re.compile("<!\-\-[\s\S]+?\-\->")
def rmtags(txt):
    txt = rerem.sub(' ', txt)
    a = retag.split(txt)
    b = ''.join(a)
    return b.replace("&nbsp;", " ").replace("&ndash;", "-")

def wordtok(tks, junk = re.compile('[\d\.!\\:;\/\?,\(\)\[\]"]+')):
    return filter(lambda x: not junk.match(x), tks)

def fd(ws):
    d = {}
    for w in ws:
        d[w] = d.get(w, 0) + 1
    return d

def fdall(chunks):
    d = {}
    for ws in chunks:
        for w in ws:
            d[w] = d.get(w, 0) + 1
    return d

def csv(xs, fn):
    return sav(''.join(["\t".join([str(y) for y in x]) + "\n" for x in xs]), fn)

import random
def argm(d, f = max):
    if len(d) == 0:
        return None
    m = f(d.values())
    return random.choice([x for x in d.keys() if d[x] == m])

def allargm(d, f = max):
    m = f(d.values())
    return [x for (x, v) in d.items() if v == m]

def amax(d, f=max):
    return f((v, k) for (k, v) in d.items())


#pd = norm(fd( whatever ))
#iv = ival(pd)
#pick(iv)
def ival(p):
    a = 0.0
    r = {}
    for (k, v) in p.items():
        r[k] = a + v
        a += v
    return r

def pick(iv):
    a = random.random()
    (v, x) = min([(v, k) for (k, v) in iv.items() if a <= v])
    return x

latextemplate = """
\\documentclass[12pt]{article}
\\usepackage{rotating}
\\usepackage{fontspec}
\\usepackage{tikz-qtree}
\\usepackage[authoryear,round,comma,sort]{natbib}
\\usepackage{url}
\\usepackage{supertabular}

\\begin{document}
%s
\\bibliographystyle{apalike}
\\bibliography{hh,hhling}
\\end{document}
"""

latextemplateu = u"""
\\documentclass[12pt]{article}
\\usepackage[T3,T1]{fontenc}
\\usepackage[noenc]{tipa}
\\usepackage{lmodern}
\\usepackage[authoryear,round,comma,sort]{natbib}
\\usepackage{hyperref}
\\usepackage{bibentry}
\\usepackage[xetex]{graphicx}
\\usepackage{alltt}
\\usepackage{supertabular}
\\usepackage{longtable}
\\usepackage{microtype}
\\usepackage[combine,tipa]{ucs}
\\setcitestyle{authoryear,round,aysep={},yysep={,},notesep={:},semicolon}

\\begin{document}
%s
\\bibliographystyle{apalike}
\\bibliography{hh,hhling}
\\end{document}
"""

def quicktabular(lraw, tabular = "tabular", c = 1, decimals = 2, sep = " ", hline = ""):
    def cell(x):
        if type(x) == type(0.0):
            return ("%." + str(decimals) + "f") % x
        return str(x)

    cells = nsplit(lraw, c)
    clist = [cells[k] for k in sorted(cells.keys())]
    l = [reduce(lambda x, y: x + y, row) for row in zip(*clist)]
    wid = max([len(x) for x in l])
    r =  ''.join([' & '.join([cell(c) for c in row]) + "\\\\" + ("\\hline" if i==0 else "") + (hline if i != len(l)-1 else "") + "\n" for (i, row) in enumerate(l)])
    return ("\\begin{%s}{l|%s}\n" % (tabular, sep.join(['r' for i in range(wid-1)]))) + r + "\\end{%s}\n" % tabular

def nsplit(l, n, default = ('', '')):
    r = {}
    coll = len(l) // n
    if len(l) % n != 0:
        coll = coll + 1
    for (i, x) in enumerate(l):
        cell = i // coll
        r[cell] = r.get(cell, []) + [x]
    for (cell, col) in r.items():
        if len(col) < coll:
            r[cell] = r[cell] + [default for _ in range(coll-len(col))]
    return r

def quicktabularu(lraw, tabular = u"tabular", c = 1, sep = u"|", rubric = u"\\hline"):
    def cell(x):
        if type(x) == type(0.0):
            return u"%.2f" % x
        return u"%s" % x

    cells = nsplit(lraw, c)
    clist = [cells[k] for k in sorted(cells.keys())]
    l = [reduce(lambda x, y: x + y, row) for row in zip(*clist)]
    wid = max([len(x) for x in l])
    r =  u''.join([u' & '.join([cell(c) for c in row]) + u"\\\\" + (rubric if i==0 else u"") + u"\n" for (i, row) in enumerate(l)])
    return (u"\\begin{%s}{l%s%s}\n" % (tabular, sep, ' '.join([u'r' for i in range(wid-1)]))) + r + u"\\end{%s}\n" % tabular

def dictmatrix_tabular(ks, output = quicktabular):
    kall = sorted(ks.keys())    
    rows = [(k1,) + tuple(["%.3f" % ks[k1][k2] for k2 in kall]) for k1 in kall]
    return output([("   ",) + tuple(kall)] + rows)

def matrix_tabular(d, output = quicktabular):
    ks = sorted(setall(d.keys()))
    toprow = [""] + [str(k) for k in ks]
    rows = [[k1] + [d.get((k1, k2), "-") for k2 in ks] for k1 in ks]
    return output([toprow] + rows)

def htmltabular(l):
    def cell(x):
        if type(x) == type(0.0):
            return "%.2f" % x
        return str(x)

    r =  ''.join(["<tr>" + ''.join(["<td>%s</td>" % cell(c) for c in row]) + "</tr>\n" for row in l])
    return "<table>\n" + r + "\n</table>\n"

import os
exts = ['zip', 'pdf', 'doc', 'djvu', 'bib', 'html', 'txt', 'tsv', 'csv', 'xls', 'xlsx', 'tab']
reext = "(?:" + '|'.join(["(?:" + z + ")" for z in exts + [z.upper() for z in exts]]) + ")"
#[zpdbZPDB][idoIDO][pfcbFPCB
rev2 = re.compile("(v\d+)?((?:\_o)?\.%s)" % reext)
def incv(s):
    def ivh(o):
        if o.group(1):
            return "v" + str(int(o.group(1)[1:]) + 1) + o.group(2)
        else:
            return "v2" + o.group(2)
    return rev2.sub(ivh, s)

def bak(fn, ext = 'old'):
    if not os.path.exists(fn):
        return False 
    thislen = os.path.getsize(fn)
    newf = takeuntil(fn, '.') + ext + "." + takeafter(fn, '.')
    while os.path.exists(newf):
        if thislen == os.path.getsize(newf):
            print(newf, "not saved in since", fn, "has same size")
            return False
        newf = incv(newf)
    os.system("copy " + fn + " " + newf)
    print(newf, "saved")
    return True

def allmax(d, f=max):
    if not d:
        return {}
    (v, _) = amax(d, f=f)
    return {k: dk for (k, dk) in d.items() if dk == v}

def frange(start, end=None, inc=None):
    "A range function, that does accept float increments..."

    if end == None:
        end = start + 0.0
        start = 0.0

    if inc == None:
        inc = 1.0

    L = []
    while 1:
        next = start + len(L) * inc
        if inc > 0 and next >= end:
            break
        elif inc < 0 and next <= end:
            break
        L.append(next)
        
    return L

htmlclicktable = """
<html>
<head>
<meta http-equiv="Content-Type" content="text/html;">
<meta name="robots" content="noindex">
<meta name="googlebot" content="noindex">
<meta http-equiv="Pragma" content="no-cache">
<title>Clickable Table</title>
<style type="text/css">
</style>
<script type="text/javascript">
<!--
function switchMenu(obj) {
   var el = document.getElementById(obj);
   if ( el.style.display != "none" ) {
      el.style.display = 'none';
   }
   else {
      el.style.display = '';
   }
}
//-->
</script>
<style type="text/css">
tr.d0 td {
  background-color: #CC9999; color: black;
}
tr.d1 td {
  background-color: #9999CC; color: black;
}
</style>
</head>
<body>
%s
</body>
</html>
"""

def downloadlink(f, title = "Download"):
    return '<a href="data:text/plain;base64,%s">%s</a>' % (f.encode("base64").replace("\n", ""), title)


reokkey = re.compile("[^A-Za-z\d\-\_\[\]]")
acwrap = """<a onclick="switchMenu('%s');" style=color:blue;cursor:pointer;>%s</a>"""
vwrap = """<div id="%s" style=display:none;>%s</div>"""

def showclick(linkname, content, cid):
    vid = reokkey.sub("_", cid)
    hst = vwrap % (vid, content)
    return (acwrap % (vid, linkname)) + hst


def clickcell(vhist, cid, cellstyle = None):
    (v, hist) = vhist
    vid = reokkey.sub("_", cid)
    hst = vwrap % (vid, '<table>%s</table>' % htmlhist(hist))
    if type(v) == type(0.0):
        if cellstyle == "percent":
            v = ("%.1f" % (100*v)) + "&#37;"
        else:
            v = "%.3f" % v
    return (acwrap % (vid, v)) + hst

def infirsttag(txt):
    txt = rerem.sub(' ', txt)
    a = retag.split(txt)
    b = a[0]
    return b.replace("&nbsp;", " ")

def tabtablehtml(rows):
    tab = ''.join(['\t'.join([infirsttag(str(x)) for x in row]) + '\n' for row in rows])
    return tab

def tabtable(rows):
    try:
        tab = ''.join(['\t'.join([str(x) for x in row]) + '\n' for row in rows])
    except UnicodeEncodeError:
        print(x, row)
        raise UnicodeEncodeError
    return tab

def uc(x):
    if type(x) == type(u""):
        return x
    if type(x) == type(""):
        return unicode(x, "ISO-8859-1")
    return u"%s" % x
    

def tabtxt(rows):
    tab = u''.join([u'\t'.join([uc(x) for x in row]) + u'\n' for row in rows])
    return tab


def htmlhist(hist):
    def shvs(vs):
        if type(vs) == type(()):
            return u"".join([shv(v) for v in vs])
        return shv(vs)
    
    def shv(v):
        if type(v) == type(0.0):
            return "%.3f" % v
        return u"<td>%s</td>" % v

    
    return ''.join(['<tr><td><FONT COLOR="green">%s</FONT></td>%s</tr>' % (k, shvs(v)) for (k, v) in hist])


searchtablejs = """
<link rel="stylesheet" href="https://cdn.datatables.net/1.10.12/css/jquery.dataTables.css">
	<script type="text/javascript" language="javascript" src="https://code.jquery.com/jquery-1.12.3.min.js">
	</script>
	<script type="text/javascript" language="javascript" src="https://cdn.datatables.net/1.10.12/js/jquery.dataTables.min.js">
	</script>
	<script type="text/javascript" class="init">
$(document).ready(function() {
	$('table.display tfoot th').each( function () {
		var title = $(this).text();
		$(this).html( '<input type="text" placeholder="Search '+title+'" />' );
	} );


        $('table.display').each( function () {
            var table = $(this).DataTable();
	    table.columns().every( function () {
		var that = this;

		$( 'input', this.footer() ).on( 'keyup change', function () {
			if ( that.search() !== this.value ) {
				that
					.search( this.value )
					.draw();
			}
		} );
	} );
} );

        });

	</script>
"""

def mediawikitable(rows, toprow = None):
    if toprow:
        tr = u'|%s|\n' % u' | '.join(toprow)
        table = tr + u"".join([c if c in ["|", "\n"] else "-" for c in tr]).replace("-|-", " | ")
    else:
        table = ""
    table = table + u''.join([u'|%s|\n' % u' | '.join(row) for row in rows])
    return table

def parsemediawiki(txt):
    rows = [row.split("|") for row in txt.split("\n") if row.strip()]
    return [tuple([x.strip() for x in r[1:-1]]) for r in rows]

def searchtable(rows, toprow, align = "right", tableid = "tableid"):
    tableopts = (' id="%s"' % tableid) + ' class="display" cellspacing="0" width="100%"'
    return htmltable(rows, toprow = toprow, align = align, tableopts = tableopts)

def htmltable(rows, toprow = None, download = "", align = "right", idprefix = "", exceldownload = False, tableopts = "", cellstyle = None):
    def tdhist(td, cid = ""):
        #print type(td), td[:2]
        if type(td) == type(()):
           return clickcell(td, cid, cellstyle = cellstyle)
        return u"%s" % td

    if toprow:
        table = u"<thead><tr>" + u''.join([u'<th align="center" colspan="%s">%s</th>' % td if type(td) == type(()) else u'<th align="center">%s</th>' % td for td in toprow]) + u"</tr></thead>"
        if tableopts:
            table = table + u"<tfoot><tr>" + u''.join([u'<th align="center">%s</th>' % td for td in toprow]) + u"</tr></tfoot>"
        allrows = [toprow] + rows
    else:
        table = ""
        allrows = rows
    table = table + u"<tbody>" + u''.join([(u'<tr class="d%s">' % (i % 2)) + u"<td><b>" + tdhist(tr[0], u"%s%s_%s" % (idprefix, i, 0)) + u"</b></td>" + ''.join([(u'<td align="%s">' % align) + tdhist(td, u"%s%s_%s" % (idprefix, i, j+1)) + u"</td>" for (j, td) in enumerate(tr[1:])]) + u"</tr>" for (i, tr) in enumerate(rows)]) + u"</tbody>"

    download = ((downloadlink(tabtablehtml(allrows), u"Export this table to Excel") + u" ") if exceldownload else u"") + download

    return ((u"<p>%s</p>" % download) if download else "") + (u"<table%s>\n" % tableopts) + table + u"\n</table>"

#(Aikana, (s1, hist))

#def htmlcolors(fn = "htmlcolors.txt"):
#    return dict([(x[0], x[3][1:]) for x in ptab(fn, i=0) if "".join(x).strip()])

retables = re.compile('\<[Tt][Aa][Bb][Ll][Ee][^\>]*\>([\s\S]*?)\<.[Tt][Aa][Bb][Ll][Ee]\>')
def htmltables(txt):
    return retables.findall(txt)

rerows = re.compile('\<[Tt][Rr][^\>]*\>([\s\S]*?)\<.[Tt][Rr]\>')
#Input everything between table /table
recells = re.compile('\<[Tt][DdHh][^\>]*\>([\s\S]*?)\<.[Tt][DdHh]\>')
def htmltabletom(x):
    rows = rerows.findall(x)
    rowcells = [recells.findall(row) for row in rows]
    return [[rmtags(c).strip() for c in row] for row in rowcells]


def delta1all(a, b, case_sensitive = False):
    if a and b:
        if not case_sensitive:
            a = a.lower()
            b = b.lower()
        if a == b:
            return 0
        else:
            return 1
    else:
        return 1


def delta(a, b, case_sensitive = False, inscost = 1, subcost = 0.5):
    if a and b:
        if not case_sensitive:
            a = a.lower()
            b = b.lower()
        if a == b:
            return 0
        else:
            return subcost
    else:
        return inscost


def deltaw(a, b, case_sensitive = False):
    if a and b:
        (md, align) = edist(a, b, delta, case_sensitive = case_sensitive)
        return md
    return 1.0

def align(d, x, y):
    i = len(x)
    j = len(y)
    r = {}
    k = 0
    #print d
    while (0, 0) != (i, j):
        #print i, j
        (_, t) = d[i, j]
        if t == 'S':
            r[k] = (x[i-1], y[j-1])
            i = i - 1
            j = j - 1
        elif t == 'I':
            r[k] = ("", y[j-1]) #'-'.ljust(len(y[j-1]))
            j = j - 1
        elif t == 'D':
            r[k] = (x[i-1], "") #'-'.ljust(len(x[i-1]))
            i = i - 1
        k = k + 1
    return [r[k] for k in reversed(range(len(r)))] #sorted(r, reverse=True)]
        
def edist(x, y, delta = delta, case_sensitive = False):
    lx = len(x)
    ly = len(y)

    d = {}
    d[(0,0)] = (0, 'None')
    for i in range(lx):
        d[(i+1, 0)] = (d[(i, 0)][0] + delta(x[i], None, case_sensitive = case_sensitive), 'D')
    for j in range(ly):
        d[(0, j+1)] = (d[(0, j)][0] + delta(None, y[j], case_sensitive = case_sensitive), 'I')

    for i in range(lx):
        for j in range(ly):
            d[(i+1, j+1)] = min((d[(i, j)][0] + delta(x[i], y[j], case_sensitive = case_sensitive), 'S'),
                                (d[(i, j+1)][0] + delta(x[i], None, case_sensitive = case_sensitive), 'D'),
                                (d[(i+1, j)][0] + delta(None, y[j], case_sensitive = case_sensitive), 'I'))

    nrm = float(max(lx, ly))
    (md, _) = d[(lx, ly)]
    return (md/nrm, align(d, x, y))


def closest(x, ys):
    return argm(dict([(y, edist(x, y)) for y in ys]), min)


def binvalues(xs, binsize = 0.05):
    bins = {}
    xss = list(sorted(xs))
    i = 0
    lim = xss[0] + binsize if type(xss[0]) == type(float) else (xss[0][0] + binsize, xss[0][1])
    while i < len(xss):
        if xss[i] <= lim:
            bins[lim] = bins.get(lim, []) + [xss[i]]
            i = i + 1
        else: 
            lim = lim + binsize if (type(lim) == type(float)) else (lim[0] + binsize, lim[1])
    return bins 


import subprocess
def ossystem(s, shell = False):
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startupinfo.wShowWindow = subprocess.SW_HIDE
    return subprocess.call(s, startupinfo=startupinfo, shell = shell) 

